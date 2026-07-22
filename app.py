import json, os, secrets
from datetime import datetime, date, timedelta, timezone
from io import BytesIO
from zoneinfo import ZoneInfo
from functools import wraps
from pathlib import Path

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, Response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect, text, case, or_
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from jinja2 import DictLoader

BASE = Path(__file__).resolve().parent
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
db_url = os.environ.get('DATABASE_URL', f"sqlite:///{BASE/'rimel.db'}")
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024
db = SQLAlchemy(app)

EMBEDDED_TEMPLATES = {'base.html': '<!doctype html><html lang="es"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><meta name="theme-color" content="#0876a8"><link rel="manifest" href="/static/manifest.json"><style>:root{--blue:#0876a8;--dark:#202b38;--bg:#f4f7fa;--line:#dbe3ea}*{box-sizing:border-box}body{margin:0;background:var(--bg);font:15px system-ui,-apple-system,Segoe UI,sans-serif;color:var(--dark)}header{position:sticky;top:0;z-index:10;display:flex;align-items:center;gap:14px;padding:10px 22px;background:#fff;border-bottom:1px solid var(--line)}header img{width:58px;height:48px;object-fit:contain}header div{display:flex;flex-direction:column}header small{color:#64748b}nav{margin-left:auto;display:flex;gap:8px;flex-wrap:wrap}nav a{padding:8px 10px;text-decoration:none;color:var(--dark);border-radius:8px}nav a:hover{background:#e8f4f9}main{max-width:1450px;margin:auto;padding:22px}h1{margin:.2em 0}h2{margin-top:0}.muted{color:#64748b}.flash{padding:12px;border-radius:10px;margin-bottom:12px}.flash.error{background:#fee2e2}.flash.success{background:#dcfce7}.login-card,.panel{background:#fff;border:1px solid var(--line);border-radius:14px;padding:20px;box-shadow:0 4px 16px #1e293b0c}.login-card{max-width:420px;margin:8vh auto}.narrow{max-width:520px;margin:auto}form{display:grid;gap:12px}label{display:grid;gap:6px;font-weight:600}input,select,textarea,button{font:inherit;padding:11px;border:1px solid #cbd5e1;border-radius:9px}button{background:var(--blue);color:white;border:0;font-weight:700;cursor:pointer}.secondary{background:#e7f3f8;color:#075b80}.filters{display:flex;gap:10px;flex-wrap:wrap;margin:14px 0}.filters select{min-width:180px}.metrics{display:grid;grid-template-columns:repeat(8,minmax(135px,1fr));gap:12px;margin:16px 0}.metrics article{background:#fff;border:1px solid var(--line);border-radius:12px;padding:15px}.metrics small{display:block;color:#64748b}.metrics strong{font-size:26px}.table-wrap{max-height:60vh;overflow:auto;border:1px solid var(--line);border-radius:10px}table{border-collapse:collapse;width:100%;background:#fff}th,td{text-align:left;padding:11px;border-bottom:1px solid var(--line);vertical-align:top}thead th{position:sticky;top:0;background:#eef5f8;z-index:2}td small{display:block;color:#64748b}.topline{display:flex;justify-content:space-between;align-items:center}.pill{background:#dcfce7;padding:7px 11px;border-radius:999px}.grid2{display:grid;grid-template-columns:1fr 2fr;gap:16px}.inline{display:flex;gap:6px;align-items:center}dialog{width:min(620px,94vw);border:0;border-radius:16px;padding:0;box-shadow:0 24px 80px #0005}dialog form{padding:20px}.dialog-head{display:flex;justify-content:space-between}.x{background:transparent;color:#111;font-size:28px;padding:0}.actions{display:flex;justify-content:flex-end;gap:8px}@media(max-width:900px){header{align-items:flex-start;flex-wrap:wrap}nav{width:100%;overflow:auto;flex-wrap:nowrap}.metrics{grid-template-columns:repeat(2,1fr)}.grid2{grid-template-columns:1fr}main{padding:12px}th,td{min-width:110px}.table-wrap{max-height:65vh}}\n.button{display:inline-flex;align-items:center;justify-content:center;padding:10px 12px;border-radius:9px;text-decoration:none;background:var(--blue);color:#fff;font-weight:700}.danger{background:#b42318}.compact{display:flex;align-items:center;gap:6px;font-size:13px}.compact input{min-width:145px}.section-head{display:flex;align-items:center;justify-content:space-between;gap:12px}.row-actions{display:flex;gap:7px;flex-wrap:wrap}.gps-control{display:flex;align-items:center;justify-content:space-between;gap:20px;margin:14px 0}.gps-control h2{margin-bottom:4px}.live-map{height:58vh;min-height:420px;border-radius:10px}.small-table{max-height:42vh}.status{display:inline-block;padding:5px 9px;border-radius:999px;font-size:12px;font-weight:700}.status.online{background:#dcfce7;color:#166534}.status.offline{background:#f1f5f9;color:#475569}@media(max-width:900px){.gps-control,.section-head{align-items:stretch;flex-direction:column}.live-map{height:52vh;min-height:340px}.filters .compact{width:100%;justify-content:space-between}.filters .compact input{flex:1}}\n.client-form{display:grid;grid-template-columns:1fr 1.2fr 1fr 1.6fr;gap:12px;align-items:end}.client-form .wide{grid-column:span 2}.client-form label,.optimizer-grid label,.stack-form label{display:flex;flex-direction:column;gap:5px}.optimizer-grid{display:grid;grid-template-columns:1fr 1fr 2fr 1fr auto;gap:12px;align-items:end}.optimizer-grid .wide{min-width:250px}.progress{height:10px;background:#e6edf1;border-radius:999px;overflow:hidden;margin:14px 0}.progress span{display:block;height:100%;width:0;background:#0876a8;transition:width .2s}.route-preview{max-height:380px;overflow:auto}.route-preview li{margin:6px 0}.inactive-row{opacity:.65;background:#f3f3f3}.warning-text{color:#a85d00}.narrow-panel{max-width:720px}.stack-form{display:grid;gap:14px}.row-actions form{display:inline}.row-actions form button{height:100%}@media(max-width:900px){.client-form,.optimizer-grid{grid-template-columns:1fr 1fr}.client-form .wide,.optimizer-grid .wide{grid-column:span 2}}@media(max-width:600px){.client-form,.optimizer-grid{grid-template-columns:1fr}.client-form .wide,.optimizer-grid .wide{grid-column:span 1}}\n/* v2026.07.20.6: mapa Leaflet y exportaciones */\n.excel{background:#147d3f}.map-panel{padding:0;overflow:hidden}.live-map{position:relative;width:100%;height:560px;min-height:420px;overflow:hidden;background:#e5e7eb;isolation:isolate}.leaflet-container{font:inherit;z-index:1}.leaflet-container img,.leaflet-tile,.leaflet-marker-icon,.leaflet-marker-shadow{max-width:none!important;max-height:none!important;width:auto;height:auto}.leaflet-tile-pane img{position:absolute!important}.complaint-form{display:grid;grid-template-columns:repeat(3,minmax(120px,1fr));gap:6px;min-width:520px}.complaint-form textarea{grid-column:1/-1}.complaint-form button{grid-column:1/-1}.version{font-size:11px;color:#0876a8;font-weight:700}.actions{flex-wrap:wrap}@media(max-width:900px){.live-map{height:55vh;min-height:360px}.complaint-form{min-width:420px;grid-template-columns:1fr}.complaint-form textarea,.complaint-form button{grid-column:auto}}\n\n.daily-route-map{height:560px;min-height:390px;width:100%;border-radius:12px;background:#e5e7eb}.route-map-legend{display:flex;gap:12px;flex-wrap:wrap;margin:10px 0}.route-map-legend span{display:inline-flex;align-items:center;gap:6px}.route-number-icon,.origin-icon{background:transparent!important;border:0!important}.route-pin{position:relative;width:44px;height:54px;filter:drop-shadow(0 4px 5px #0007)}.route-pin-shape{position:absolute;left:3px;top:2px;width:38px;height:38px;background:#0876a8;border:3px solid #fff;border-radius:50% 50% 50% 0;transform:rotate(-45deg)}.route-pin-number{position:absolute;left:3px;top:2px;width:38px;height:38px;display:flex;align-items:center;justify-content:center;color:#fff;font-size:15px;font-weight:900;line-height:1;z-index:2}.origin-pin{position:relative;width:48px;height:58px;filter:drop-shadow(0 4px 5px #0007)}.origin-pin-shape{position:absolute;left:3px;top:2px;width:42px;height:42px;background:#dc2626;border:3px solid #fff;border-radius:50% 50% 50% 0;transform:rotate(-45deg)}.origin-pin-label{position:absolute;left:3px;top:2px;width:42px;height:42px;display:flex;align-items:center;justify-content:center;color:#fff;font-size:18px;font-weight:900;z-index:2}.client-popup{min-width:230px}.client-popup h3{margin:0 0 6px;font-size:16px;color:#075b80}.client-popup .popup-code{display:inline-block;background:#e7f3f8;color:#075b80;padding:3px 7px;border-radius:999px;font-size:11px;font-weight:800;margin-bottom:7px}.client-popup p{margin:4px 0}.client-popup .popup-actions{display:flex;gap:6px;flex-wrap:wrap;margin-top:10px}.client-popup .popup-actions a{display:inline-block;text-decoration:none;padding:7px 9px;border-radius:7px;background:#0876a8;color:#fff;font-weight:700;font-size:12px}.client-popup .popup-actions a.secondary{background:#e7f3f8;color:#075b80}@media(max-width:900px){.daily-route-map{height:62vh;min-height:390px}}\n.release-banner{background:#0f766e;color:#fff;text-align:center;padding:7px 12px;font-weight:800;letter-spacing:.02em}.leaflet-container .leaflet-pane,.leaflet-container .leaflet-map-pane,.leaflet-container .leaflet-tile-pane{position:absolute;left:0;top:0}.leaflet-container .leaflet-tile{position:absolute!important;width:256px!important;height:256px!important;max-width:none!important;max-height:none!important}.leaflet-container{overflow:hidden!important}\n</style><title>{% block title %}RIMEL{% endblock %}</title></head><body>\n<header><img src="/static/logo.png" alt="RIMEL"><div><strong>Control de Vendedores</strong><small>{% if current_user %}{{ current_user.full_name }} · {{ \'Gerencial\' if current_user.role==\'manager\' else \'Vendedor\' }}{% endif %}</small><small class="version">v{{app_version}}</small></div>{% if current_user %}<nav><a href="/">Inicio</a><a href="/history">Historial</a><a href="/complaints">Quejas</a>{% if current_user.role==\'manager\' %}<a href="/tracking">GPS en vivo</a><a href="/clients">Clientes y rutas</a><a href="/users">Usuarios</a>{% endif %}<a href="/change-password">Contraseña</a><a href="/logout">Salir</a></nav>{% endif %}</header>\n<div class="release-banner">ACTUALIZACIÓN ACTIVA · v{{app_version}} · Hora Rivera</div><main>{% with messages=get_flashed_messages(with_categories=true) %}{% for cat,msg in messages %}<div class="flash {{cat}}">{{msg}}</div>{% endfor %}{% endwith %}{% block content %}{% endblock %}</main>\n<script>if(\'serviceWorker\' in navigator){navigator.serviceWorker.getRegistrations().then(rs=>rs.forEach(r=>r.unregister()));}if(window.caches){caches.keys().then(keys=>keys.forEach(k=>caches.delete(k)));}</script>{% block scripts %}{% endblock %}</body></html>\n', 'change_password.html': '{% extends \'base.html\' %}{% block content %}<section class="panel narrow"><h1>Cambiar contraseña</h1><form method="post"><label>Nueva contraseña<input type="password" name="password" minlength="8" required></label><label>Confirmar<input type="password" name="confirm" minlength="8" required></label><button>Guardar</button></form></section>{% endblock %}\n', 'client_edit.html': '{% extends \'base.html\' %}\n{% block title %}Modificar cliente · RIMEL{% endblock %}\n{% block content %}\n<div class="topline"><div><h1>Modificar cliente</h1><p class="muted">Solo Gerencia puede cambiar estos datos.</p></div><a class="button secondary" href="/clients">Volver</a></div>\n<section class="panel narrow-panel"><form class="stack-form" method="post">\n<label>Código<input name="code" value="{{client.code}}" required></label>\n<label>Nombre<input name="name" value="{{client.name}}" required></label>\n<label>Dirección<input name="address" value="{{client.address}}" required></label>\n<label>Vendedor<input name="seller" list="sellerList" value="{{client.seller}}" required></label>\n<datalist id="sellerList">{% for s in sellers %}<option value="{{s}}">{% endfor %}</datalist>\n<label>Día<select name="day">{% for d in [\'LUNES\',\'MARTES\',\'MIERCOLES\',\'JUEVES\',\'VIERNES\',\'SABADO\',\'DOMINGO\'] %}<option {{\'selected\' if client.day==d}}>{{d}}</option>{% endfor %}</select></label>\n<div class="actions"><button>Guardar cambios</button><a class="button secondary" target="_blank" href="/client/{{client.id}}/map">Ver mapa</a></div>\n</form></section>\n{% endblock %}\n', 'clients.html': '{% extends \'base.html\' %}\n{% block title %}Clientes y rutas · RIMEL{% endblock %}\n{% block content %}\n<div class="topline"><div><h1>Clientes y rutas</h1><p class="muted">Administración exclusiva de Gerencia. Aquí puedes crear, modificar, reasignar, dar de baja y optimizar rutas.</p></div><span class="pill">{{clients|length}} registros</span></div>\n\n<section class="panel">\n  <h2>Agregar cliente</h2>\n  <form class="client-form" method="post" action="/clients/create">\n    <label>Código<input name="code" placeholder="Automático si se deja vacío"></label>\n    <label>Vendedor<input name="seller" list="sellerList" required></label>\n    <label>Día<select name="day" required>{% for d in [\'LUNES\',\'MARTES\',\'MIERCOLES\',\'JUEVES\',\'VIERNES\',\'SABADO\',\'DOMINGO\'] %}<option>{{d}}</option>{% endfor %}</select></label>\n    <label>Nombre<input name="name" required></label>\n    <label class="wide">Dirección<input name="address" required placeholder="Calle y número, Rivera"></label>\n    <button>Agregar cliente</button>\n  </form>\n  <datalist id="sellerList">{% for s in sellers %}<option value="{{s}}">{% endfor %}</datalist>\n</section>\n\n<section class="panel optimizer-box">\n  <div class="section-head"><div><h2>Optimización del recorrido</h2><p class="muted">La secuencia se guarda en la base central. Desde ese momento, el vendedor verá la ruta ya ordenada al seleccionar su día.</p></div></div>\n  <div class="optimizer-grid">\n    <label>Vendedor<select id="optSeller"><option value="">Seleccionar</option>{% for s in sellers %}<option>{{s}}</option>{% endfor %}</select></label>\n    <label>Día<select id="optDay"><option value="">Seleccionar</option>{% for d in [\'LUNES\',\'MARTES\',\'MIERCOLES\',\'JUEVES\',\'VIERNES\',\'SABADO\',\'DOMINGO\'] %}<option>{{d}}</option>{% endfor %}</select></label>\n    <label class="wide">Punto de partida<input id="optStart" value="Sarandí 1700, Shopping Melancia, Rivera, Uruguay" readonly></label>\n    <label>Final<select id="optReturn"><option value="open">Último cliente</option><option value="return">Regresar al origen</option></select></label>\n    <button type="button" id="optimizeBtn">Optimizar ruta seleccionada</button><button type="button" class="secondary" id="optimizeAllBtn">Optimizar todas las rutas</button>\n  </div>\n  <div class="progress"><span id="optProgress"></span></div>\n  <p id="optStatus" class="muted">Selecciona vendedor y día. La primera geocodificación puede demorar algunos minutos.</p>\n  <div id="optPreview" class="route-preview"></div>\n</section>\n\n<form class="filters" method="get">\n  <select name="seller"><option value="">Todos los vendedores</option>{% for s in sellers %}<option value="{{s}}" {{\'selected\' if s==selected_seller}}>{{s}}</option>{% endfor %}</select>\n  <select name="day"><option value="">Todos los días</option>{% for d in [\'LUNES\',\'MARTES\',\'MIERCOLES\',\'JUEVES\',\'VIERNES\',\'SABADO\',\'DOMINGO\'] %}<option value="{{d}}" {{\'selected\' if d==selected_day}}>{{d}}</option>{% endfor %}</select>\n  <select name="status"><option value="active" {{\'selected\' if selected_status==\'active\'}}>Activos</option><option value="inactive" {{\'selected\' if selected_status==\'inactive\'}}>Inactivos</option><option value="all" {{\'selected\' if selected_status==\'all\'}}>Todos</option></select>\n  <input name="q" value="{{search}}" placeholder="Buscar cliente, código o dirección">\n  <button>Filtrar</button><a class="button secondary" href="/clients">Limpiar</a>\n</form>\n\n<section class="panel"><div class="table-wrap"><table><thead><tr><th>Orden</th><th>Código</th><th>Cliente</th><th>Dirección</th><th>Vendedor</th><th>Día</th><th>Estado</th><th>Acciones</th></tr></thead><tbody>\n{% for c in clients %}<tr class="{{\'inactive-row\' if not c.active}}"><td>{{c.optimized_order or c.original_order}}</td><td>{{c.code}}</td><td><b>{{c.name}}</b></td><td>{{c.address}}</td><td>{{c.seller}}</td><td>{{c.day}}</td><td><span class="pill">{{\'Activo\' if c.active else \'Inactivo\'}}</span></td><td><div class="row-actions"><a class="button secondary" href="/clients/{{c.id}}/edit">Modificar</a><a class="button secondary" target="_blank" href="/client/{{c.id}}/map">Mapa</a><form method="post" action="/clients/{{c.id}}/toggle" onsubmit="return confirm(\'{{\'Dar de baja\' if c.active else \'Reactivar\'}} este cliente?\')"><button class="{{\'danger\' if c.active else \'\'}}">{{\'Dar de baja\' if c.active else \'Reactivar\'}}</button></form></div></td></tr>\n{% else %}<tr><td colspan="8">No hay clientes para estos filtros.</td></tr>{% endfor %}\n</tbody></table></div></section>\n{% endblock %}\n{% block scripts %}\n<script>\nconst ORIGIN={latitude:-30.9173162,longitude:-55.5488927,name:\'Sarandí 1700 · Shopping Melancia\'};\nconst sleep=ms=>new Promise(r=>setTimeout(r,ms));\nconst statusEl=document.getElementById(\'optStatus\'),progressEl=document.getElementById(\'optProgress\'),previewEl=document.getElementById(\'optPreview\');\nfunction hav(a,b){const R=6371,d2r=Math.PI/180,dLat=(b.latitude-a.latitude)*d2r,dLon=(b.longitude-a.longitude)*d2r;const x=Math.sin(dLat/2)**2+Math.cos(a.latitude*d2r)*Math.cos(b.latitude*d2r)*Math.sin(dLon/2)**2;return 2*R*Math.asin(Math.sqrt(x));}\nfunction cleanAddress(a){return String(a||\'\').replace(/\\bEsq\\.?\\s*:/gi,\' esquina \').replace(/\\bApto\\.?\\s*:/gi,\' \').replace(/\\bFte\\.?/gi,\' frente a \').replace(/\\s+/g,\' \').trim();}\nasync function geocode(text){const key=\'geo:\'+text.toUpperCase();try{const saved=localStorage.getItem(key);if(saved)return JSON.parse(saved);}catch(e){}const queries=[`${cleanAddress(text)}, Rivera, Uruguay`,`${cleanAddress(text)}, 40000 Rivera, Uruguay`];for(const query of queries){const url=`https://nominatim.openstreetmap.org/search?format=jsonv2&limit=1&countrycodes=uy&accept-language=es&q=${encodeURIComponent(query)}`;const r=await fetch(url,{headers:{Accept:\'application/json\'}});if(!r.ok)continue;const data=await r.json();if(data.length){const point={latitude:Number(data[0].lat),longitude:Number(data[0].lon)};try{localStorage.setItem(key,JSON.stringify(point));}catch(e){}return point;}await sleep(1100);}return null;}\nfunction routeDistance(path,start,back){let total=0,prev=start;for(const p of path){total+=hav(prev,p);prev=p;}if(back&&path.length)total+=hav(prev,start);return total;}\nfunction nearest(points,start){const pending=[...points],path=[];let cur=start;while(pending.length){let bi=0,bd=Infinity;pending.forEach((p,i)=>{const d=hav(cur,p);if(d<bd){bd=d;bi=i;}});cur=pending.splice(bi,1)[0];path.push(cur);}return path;}\nfunction twoOpt(path,start,back){let best=[...path],bestD=routeDistance(best,start,back),changed=true,passes=0;while(changed&&passes<12){changed=false;passes++;for(let i=0;i<best.length-2;i++){for(let k=i+1;k<best.length-1;k++){const candidate=best.slice(0,i).concat(best.slice(i,k+1).reverse(),best.slice(k+1));const d=routeDistance(candidate,start,back);if(d+0.005<bestD){best=candidate;bestD=d;changed=true;}}}}return {path:best,distance:bestD,method:\'distancia geográfica\'};}\nasync function roadTrip(points,back){if(points.length>98)throw new Error(\'Ruta extensa: se usa optimización geográfica\');const coords=[ORIGIN,...points].map(p=>`${p.longitude},${p.latitude}`).join(\';\');const round=back?\'true\':\'false\';const dest=back?\'last\':\'any\';const url=`https://router.project-osrm.org/trip/v1/driving/${coords}?source=first&destination=${dest}&roundtrip=${round}&overview=false&steps=false`;\n const r=await fetch(url);if(!r.ok)throw new Error(\'Servicio vial no disponible\');const data=await r.json();if(data.code!==\'Ok\'||!data.trips?.length)throw new Error(\'No se pudo calcular por calles\');const ordered=data.waypoints.slice(1).map((w,i)=>({point:points[i],index:w.waypoint_index})).sort((a,b)=>a.index-b.index).map(x=>x.point);return {path:ordered,distance:data.trips[0].distance/1000,method:\'red vial\'};}\nasync function loadAndLocate(seller,day,onProgress){const rr=await fetch(`/api/routes/${encodeURIComponent(seller)}/${encodeURIComponent(day)}`);if(!rr.ok)throw new Error(\'No se pudo cargar la ruta\');const clients=await rr.json();if(!clients.length)return {clients,located:[],failed:[]};const located=[],failed=[];for(let i=0;i<clients.length;i++){const c=clients[i];onProgress?.(i,clients.length,c);let point=(c.latitude!=null&&c.longitude!=null)?{latitude:Number(c.latitude),longitude:Number(c.longitude)}:await geocode(c.address);if(point)located.push({...c,...point});else failed.push(c);if(c.latitude==null||c.longitude==null)await sleep(1050);}return {clients,located,failed};}\nasync function optimizeOne(seller,day,showPreview=true){statusEl.textContent=`${seller} · ${day}: cargando clientes…`;const {clients,located,failed}=await loadAndLocate(seller,day,(i,total,c)=>{progressEl.style.width=`${Math.round((i/Math.max(total,1))*75)}%`;statusEl.textContent=`${seller} · ${day}: localizando ${i+1}/${total} — ${c.name}`;});if(!clients.length)return {skipped:true};if(!located.length)throw new Error(`${seller} ${day}: no se pudo localizar ningún cliente`);statusEl.textContent=`${seller} · ${day}: calculando recorrido desde Shopping Melancia…`;let result;try{result=await roadTrip(located,false);}catch(e){result=twoOpt(nearest(located,ORIGIN),ORIGIN,false);}const ordered=result.path.concat(failed.map(c=>({...c,latitude:null,longitude:null})));const save=await fetch(\'/api/routes/save-optimized\',{method:\'POST\',headers:{\'Content-Type\':\'application/json\'},body:JSON.stringify({seller,day,clients:ordered,origin:ORIGIN})});if(!save.ok)throw new Error(`${seller} ${day}: no se pudo guardar el orden`);if(showPreview){previewEl.innerHTML=`<h3>Orden guardado desde ${ORIGIN.name}</h3><ol>`+ordered.map(c=>`<li><b>${c.name}</b> — ${c.address}${c.latitude==null?\' <span class="warning-text">(sin geocodificar; colocado al final)</span>\':\'\'}</li>`).join(\'\')+\'</ol>\';}return {count:clients.length,distance:result.distance,method:result.method,failed:failed.length};}\nasync function optimizeSelected(){const seller=document.getElementById(\'optSeller\').value,day=document.getElementById(\'optDay\').value;if(!seller||!day){statusEl.textContent=\'Selecciona vendedor y día.\';return;}document.getElementById(\'optimizeBtn\').disabled=true;document.getElementById(\'optimizeAllBtn\').disabled=true;previewEl.innerHTML=\'\';try{const r=await optimizeOne(seller,day,true);progressEl.style.width=\'100%\';statusEl.textContent=`Ruta guardada en el orden correcto: ${r.count} clientes · ${r.distance.toFixed(1)} km estimados por ${r.method} · ${r.failed} sin localizar.`;setTimeout(()=>location.href=`/clients?seller=${encodeURIComponent(seller)}&day=${day}&status=active`,1200);}catch(e){statusEl.textContent=\'Error: \'+e.message;progressEl.style.width=\'0\';}finally{document.getElementById(\'optimizeBtn\').disabled=false;document.getElementById(\'optimizeAllBtn\').disabled=false;}}\nasync function optimizeAll(){if(!confirm(\'Se optimizarán todas las rutas desde Sarandí 1700, Shopping Melancia. Puede demorar varios minutos. ¿Continuar?\'))return;const sellers=[...document.querySelectorAll(\'#optSeller option\')].map(o=>o.value).filter(Boolean),days=[\'LUNES\',\'MARTES\',\'MIERCOLES\',\'JUEVES\',\'VIERNES\',\'SABADO\',\'DOMINGO\'];const jobs=[];for(const s of sellers)for(const d of days)jobs.push([s,d]);document.getElementById(\'optimizeBtn\').disabled=true;document.getElementById(\'optimizeAllBtn\').disabled=true;previewEl.innerHTML=\'\';let done=0,totalClients=0,errors=[];try{for(const [s,d] of jobs){try{const r=await optimizeOne(s,d,false);if(!r.skipped)totalClients+=r.count||0;}catch(e){errors.push(e.message);}done++;progressEl.style.width=`${Math.round(done/jobs.length*100)}%`;await sleep(350);}statusEl.textContent=`Optimización finalizada: ${totalClients} clientes ordenados desde Shopping Melancia.${errors.length?\' Advertencias: \'+errors.length:\'\'}`;previewEl.innerHTML=\'<h3>Proceso terminado</h3><p>Los vendedores ya verán cada ruta en el orden guardado al ingresar al día correspondiente.</p>\'+(errors.length?\'<details><summary>Ver advertencias</summary><ul>\'+errors.map(x=>`<li>${x}</li>`).join(\'\')+\'</ul></details>\':\'\');}finally{document.getElementById(\'optimizeBtn\').disabled=false;document.getElementById(\'optimizeAllBtn\').disabled=false;}}\ndocument.getElementById(\'optimizeBtn\').addEventListener(\'click\',optimizeSelected);document.getElementById(\'optimizeAllBtn\').addEventListener(\'click\',optimizeAll);\n</script>\n{% endblock %}\n', 'complaints.html': '{% extends \'base.html\' %}{% block content %}\n<div class="section-head"><div><h1>Quejas y reclamos</h1><p class="muted">Seguimiento completo de cada caso, con prioridad, estado y resolución.</p></div><a class="button excel" href="/export/complaints.xlsx?{{ request.query_string.decode() }}">Exportar a Excel</a></div>\n<form class="filters" method="get">{% if current_user.role==\'manager\' %}<select name="seller"><option value="">Todos los vendedores</option>{% for s in sellers %}<option value="{{s}}" {{\'selected\' if s==selected_seller}}>{{s}}</option>{% endfor %}</select>{% endif %}<select name="status"><option value="">Todos los estados</option>{% for x in [\'PENDIENTE\',\'EN GESTION\',\'RESUELTA\'] %}<option {{\'selected\' if x==selected_status}}>{{x}}</option>{% endfor %}</select><select name="priority"><option value="">Todas las prioridades</option>{% for x in [\'BAJA\',\'MEDIA\',\'ALTA\',\'URGENTE\'] %}<option {{\'selected\' if x==selected_priority}}>{{x}}</option>{% endfor %}</select><button>Aplicar</button><a class="button secondary" href="/complaints">Limpiar</a></form>\n<section class="panel"><div class="table-wrap"><table><thead><tr><th>Fecha Rivera</th><th>Vendedor</th><th>Cliente</th><th>Queja</th><th>Gestión</th></tr></thead><tbody>{% for c in complaints %}<tr><td>{{c.created_at|rivera_dt}}</td><td>{{c.seller}}</td><td><b>{{c.client.name}}</b><small>{{c.client.code}}</small></td><td><b>{{c.category or \'OTRA\'}} · {{c.priority or \'MEDIA\'}}</b><br>{{c.text}}</td><td>{% if current_user.role==\'manager\' %}<form method="post" action="/complaints/{{c.id}}/status" class="complaint-form"><select name="category">{% for x in [\'ENTREGA\',\'PRODUCTO\',\'PRECIO\',\'ATENCION\',\'FACTURACION\',\'OTRA\'] %}<option {{\'selected\' if x==(c.category or \'OTRA\')}}>{{x}}</option>{% endfor %}</select><select name="priority">{% for x in [\'BAJA\',\'MEDIA\',\'ALTA\',\'URGENTE\'] %}<option {{\'selected\' if x==(c.priority or \'MEDIA\')}}>{{x}}</option>{% endfor %}</select><select name="status">{% for x in [\'PENDIENTE\',\'EN GESTION\',\'RESUELTA\'] %}<option {{\'selected\' if x==c.status}}>{{x}}</option>{% endfor %}</select><textarea name="resolution" rows="2" placeholder="Resolución o acción tomada">{{c.resolution or \'\'}}</textarea><button>Guardar gestión</button></form>{% else %}<b>{{c.status}}</b>{% if c.resolution %}<small>{{c.resolution}}</small>{% endif %}{% endif %}</td></tr>{% else %}<tr><td colspan="5">No hay quejas para los filtros seleccionados.</td></tr>{% endfor %}</tbody></table></div></section>\n{% endblock %}\n', 'dashboard.html': '{% extends \'base.html\' %}{% block content %}\n<div class="topline"><div><h1>{{ \'Resumen gerencial\' if current_user.role==\'manager\' else \'Mi ruta y actividad\' }}</h1><p class="muted">Información centralizada. El perfil vendedor solo visualiza su propia cartera y actividad.</p></div><span id="sync" class="pill">Sincronizado</span></div>\n<form class="filters" method="get">\n  {% if current_user.role==\'manager\' %}<select name="seller"><option value="">Todos los vendedores</option>{% for s in sellers %}<option value="{{s}}" {{\'selected\' if s==selected_seller}}>{{s}}</option>{% endfor %}</select>{% endif %}\n  <select name="day"><option value="">Todos los días</option>{% for d in [\'LUNES\',\'MARTES\',\'MIERCOLES\',\'JUEVES\',\'VIERNES\',\'SABADO\',\'DOMINGO\'] %}<option value="{{d}}" {{\'selected\' if d==selected_day}}>{{d}}</option>{% endfor %}</select>\n  <label class="compact">Desde<input type="date" name="date_from" value="{{date_from}}"></label><label class="compact">Hasta<input type="date" name="date_to" value="{{date_to}}"></label>\n  <button>Aplicar filtros</button><a class="button secondary" href="/">Limpiar</a><a class="button excel" href="/export/routes.xlsx?{{ request.query_string.decode() }}">Exportar rutas a Excel</a>\n</form>\n{% if current_user.role==\'seller\' %}\n<section class="gps-control panel"><div><h2>Seguimiento GPS de la jornada</h2><p class="muted">Actívalo al comenzar y deténlo al finalizar. Solo comparte ubicación mientras este control está activo y la aplicación permanece operativa en el teléfono.</p></div><div class="actions"><button type="button" id="startTracking">Iniciar seguimiento</button><button type="button" class="danger" id="stopTracking">Detener</button></div><p id="trackingStatus" class="muted"></p></section>\n{% endif %}\n<section class="metrics">\n{% set cards=[(\'Planificados\',\'planned\',\'\'),(\'Visitas\',\'visits\',\'\'),(\'Clientes únicos\',\'unique\',\'\'),(\'Cumplimiento\',\'compliance\',\'%\'),(\'Pedidos\',\'orders\',\'\'),(\'Conversión\',\'conversion\',\'%\'),(\'Tasa de contacto\',\'contact_rate\',\'%\'),(\'Sin reposición\',\'no_restock\',\'\'),(\'Ausentes\',\'absent\',\'\'),(\'Cerrados\',\'closed\',\'\'),(\'Reprogramados\',\'reprogrammed\',\'\'),(\'Inactivos\',\'inactive\',\'\'),(\'Registros con GPS\',\'gps_rate\',\'%\'),(\'Con observaciones\',\'notes_rate\',\'%\'),(\'Visitas repetidas\',\'repeat_visits\',\'\'),(\'Promedio diario\',\'avg_daily\',\'\'),(\'Quejas\',\'complaints\',\'\'),(\'Quejas pendientes\',\'complaints_pending\',\'\'),(\'Tasa de quejas\',\'complaint_rate\',\'%\'),(\'Vendedores activos\',\'active_sellers\',\'\')] %}\n{% for label,key,suffix in cards %}<article><small>{{label}}</small><strong>{{metrics[key]}}{{suffix}}</strong></article>{% endfor %}\n</section>\n{% if current_user.role==\'manager\' %}\n<section class="panel"><div class="section-head"><div><h2>Desempeño por vendedor</h2><p class="muted">Comparación para el período seleccionado.</p></div><a class="button" href="/tracking">Ver GPS en vivo</a></div><div class="table-wrap small-table"><table><thead><tr><th>Vendedor</th><th>Planificados</th><th>Visitas</th><th>Únicos</th><th>Cumplimiento</th><th>Pedidos</th><th>Conversión</th><th>Última ubicación</th></tr></thead><tbody>{% for s in seller_stats %}<tr><td><b>{{s.seller}}</b></td><td>{{s.planned}}</td><td>{{s.visits}}</td><td>{{s.unique}}</td><td>{{s.compliance}}%</td><td>{{s.orders}}</td><td>{{s.conversion}}%</td><td>{{s.last_seen|rivera_dt(\'%d/%m %H:%M\') if s.last_seen else \'Sin datos\'}}</td></tr>{% endfor %}</tbody></table></div></section>\n{% endif %}\n{% if current_user.role==\'seller\' and selected_day %}\n<section class="panel">\n  <div class="section-head"><div><h2>Mapa de clientes del día</h2><p class="muted">Salida desde Sarandí 1700 · Shopping Melancia. Los pines están numerados en el orden correcto del recorrido.</p></div><button type="button" class="secondary" id="fitDailyRoute">Ver ruta completa</button></div>\n  <div class="route-map-legend"><span>🔴 Inicio</span><span>📍 Pin numerado de cada cliente</span><span id="routeMapStatus" class="muted">Cargando ubicaciones…</span></div>\n  <div id="dailyRouteMap" class="daily-route-map"></div>\n</section>\n{% endif %}\n<section class="panel"><div class="section-head"><div><h2>Ruta del día en orden de visita</h2><p class="muted">La secuencia comienza en Sarandí 1700 · Shopping Melancia. Recorre los clientes de arriba hacia abajo.</p></div><span class="pill">{{clients|length}} clientes</span></div><div class="table-wrap"><table><thead><tr><th>Parada</th><th>Día</th>{% if current_user.role==\'manager\' %}<th>Vendedor</th>{% endif %}<th>Cliente</th><th>Dirección</th><th>Acciones</th></tr></thead><tbody>{% for c in clients %}<tr><td>{{c.optimized_order or c.original_order}}</td><td>{{c.day}}</td>{% if current_user.role==\'manager\' %}<td>{{c.seller}}</td>{% endif %}<td><b>{{c.name}}</b><small>{{c.code}}</small></td><td>{{c.address}}</td><td><div class="row-actions"><a class="button secondary" target="_blank" rel="noopener" href="/client/{{c.id}}/map">Mapa</a><button type="button" onclick=\'openVisit({{c.id}}, {{c.name|tojson}}, {{c.address|tojson}})\'>Registrar visita</button></div></td></tr>{% else %}<tr><td colspan="6">No hay clientes para el filtro seleccionado.</td></tr>{% endfor %}</tbody></table></div></section>\n<dialog id="visitDialog"><form method="post" action="/visits"><div class="dialog-head"><div><h2 id="vname"></h2><p id="vaddress"></p><a id="vmap" class="button secondary" target="_blank" rel="noopener">Abrir mapa</a></div><button type="button" class="x" onclick="visitDialog.close()">×</button></div><input type="hidden" name="client_id" id="client_id"><input type="hidden" name="latitude" id="latitude"><input type="hidden" name="longitude" id="longitude"><label>Resultado<select name="result" required>{% for r in allowed_results %}<option>{{r}}</option>{% endfor %}</select></label><label>Observaciones<textarea name="notes" rows="3"></textarea></label><label>Queja del cliente<textarea name="complaint" rows="3" placeholder="Registrar únicamente cuando exista una queja o reclamo"></textarea></label><div class="grid2"><label>Categoría<select name="complaint_category"><option>ENTREGA</option><option>PRODUCTO</option><option>PRECIO</option><option>ATENCION</option><option>FACTURACION</option><option selected>OTRA</option></select></label><label>Prioridad<select name="complaint_priority"><option>BAJA</option><option selected>MEDIA</option><option>ALTA</option><option>URGENTE</option></select></label></div><div class="actions"><button type="button" class="secondary" onclick="captureGPS()">Capturar GPS</button><button>Guardar y sincronizar</button></div><p id="gpsStatus" class="muted"></p></form></dialog>\n{% endblock %}{% block scripts %}\n{% if current_user.role==\'seller\' and selected_day %}\n<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">\n<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>\n<script>\nconst DAILY_ROUTE={{ route_map_clients|tojson }};\nconst ROUTE_ORIGIN={latitude:-30.9173162,longitude:-55.5488927,name:\'Sarandí 1700 · Shopping Melancia\'};\nlet dailyMap,dailyBounds=[];\nconst sleepRoute=ms=>new Promise(r=>setTimeout(r,ms));\nfunction cleanRouteAddress(a){return String(a||\'\').replace(/\\bEsq\\.?\\s*:/gi,\' esquina \').replace(/\\bApto\\.?\\s*:/gi,\' \').replace(/\\bFte\\.?/gi,\' frente a \').replace(/\\s+/g,\' \').trim();}\nasync function geocodeRouteClient(c){if(c.latitude!=null&&c.longitude!=null)return c;const queries=[`${cleanRouteAddress(c.address)}, Rivera, Uruguay`,`${cleanRouteAddress(c.address)}, 40000 Rivera, Uruguay`];for(const q of queries){try{const r=await fetch(`https://nominatim.openstreetmap.org/search?format=jsonv2&limit=1&countrycodes=uy&accept-language=es&q=${encodeURIComponent(q)}`,{headers:{Accept:\'application/json\'}});if(r.ok){const d=await r.json();if(d.length){c.latitude=Number(d[0].lat);c.longitude=Number(d[0].lon);await fetch(\'/api/clients/save-coordinate\',{method:\'POST\',headers:{\'Content-Type\':\'application/json\'},body:JSON.stringify({client_id:c.id,latitude:c.latitude,longitude:c.longitude})});return c;}}}catch(e){}await sleepRoute(1100);}return c;}\nfunction fitDaily(){if(!dailyMap)return;dailyMap.invalidateSize(true);if(dailyBounds.length)dailyMap.fitBounds(dailyBounds,{padding:[35,35],maxZoom:16});else dailyMap.setView([ROUTE_ORIGIN.latitude,ROUTE_ORIGIN.longitude],13);}\nasync function initDailyRouteMap(){dailyMap=L.map(\'dailyRouteMap\',{zoomControl:true,preferCanvas:true}).setView([ROUTE_ORIGIN.latitude,ROUTE_ORIGIN.longitude],13);L.tileLayer(\'https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png\',{subdomains:\'abc\',maxZoom:19,attribution:\'&copy; OpenStreetMap\'}).addTo(dailyMap);const oi=L.divIcon({className:\'origin-icon\',html:\'<div class=\"origin-pin\"><div class=\"origin-pin-shape\"></div><div class=\"origin-pin-label\">⌂</div></div>\',iconSize:[48,58],iconAnchor:[24,56],popupAnchor:[0,-52]});L.marker([ROUTE_ORIGIN.latitude,ROUTE_ORIGIN.longitude],{icon:oi,zIndexOffset:1000}).addTo(dailyMap).bindPopup(`<div class=\"client-popup\"><h3>Punto de partida</h3><p><b>${ROUTE_ORIGIN.name}</b></p><p>Inicio de todas las rutas del día.</p></div>`);dailyBounds=[[ROUTE_ORIGIN.latitude,ROUTE_ORIGIN.longitude]];const located=[];let missing=0;for(let i=0;i<DAILY_ROUTE.length;i++){const c=await geocodeRouteClient(DAILY_ROUTE[i]);document.getElementById(\'routeMapStatus\').textContent=`Ubicando ${i+1}/${DAILY_ROUTE.length}`;if(c.latitude!=null&&c.longitude!=null){const pos=[Number(c.latitude),Number(c.longitude)];located.push(pos);dailyBounds.push(pos);const icon=L.divIcon({className:\'route-number-icon\',html:`<div class="route-pin"><div class="route-pin-shape"></div><div class="route-pin-number">${c.order}</div></div>`,iconSize:[44,54],iconAnchor:[22,52],popupAnchor:[0,-48]});const popup=`<div class="client-popup"><h3>Parada ${c.order} · ${c.name}</h3><span class="popup-code">${c.code||\'Sin código\'}</span><p><b>Dirección:</b> ${c.address}</p><p><b>Día:</b> ${c.day||\'\'}</p><p><b>Vendedor:</b> ${c.seller||\'\'}</p><div class="popup-actions"><a target="_blank" rel="noopener" href="https://www.google.com/maps/dir/?api=1&destination=${c.latitude},${c.longitude}">Navegar</a><a class="secondary" href="/client/${c.id}/map" target="_blank" rel="noopener">Ver cliente</a></div></div>`;L.marker(pos,{icon,title:`${c.order}. ${c.name}`,riseOnHover:true}).addTo(dailyMap).bindPopup(popup,{maxWidth:320});}else missing++;}\nif(located.length)L.polyline([[ROUTE_ORIGIN.latitude,ROUTE_ORIGIN.longitude],...located],{weight:4,opacity:.7,dashArray:\'8 8\'}).addTo(dailyMap);document.getElementById(\'routeMapStatus\').textContent=`${located.length} clientes en el mapa${missing?` · ${missing} sin localizar`:\'\'}`;setTimeout(fitDaily,250);}\ndocument.addEventListener(\'DOMContentLoaded\',()=>{initDailyRouteMap();document.getElementById(\'fitDailyRoute\').onclick=fitDaily;window.addEventListener(\'resize\',()=>dailyMap&&dailyMap.invalidateSize(false));});\n</script>\n{% endif %}\n<script>\nconst dlg=document.getElementById(\'visitDialog\');\nfunction openVisit(id,n,a){client_id.value=id;vname.textContent=n;vaddress.textContent=a;vmap.href=\'/client/\'+id+\'/map\';latitude.value=\'\';longitude.value=\'\';gpsStatus.textContent=\'\';dlg.showModal()}\nfunction captureGPS(){if(!navigator.geolocation){gpsStatus.textContent=\'GPS no disponible\';return}gpsStatus.textContent=\'Obteniendo ubicación…\';navigator.geolocation.getCurrentPosition(p=>{latitude.value=p.coords.latitude;longitude.value=p.coords.longitude;gpsStatus.textContent=\'Ubicación capturada con precisión aproximada de \'+Math.round(p.coords.accuracy)+\' m\';},e=>gpsStatus.textContent=\'No fue posible obtener la ubicación: \'+e.message,{enableHighAccuracy:true,timeout:15000,maximumAge:0})}\nsetInterval(async()=>{try{const r=await fetch(\'/api/summary\');if(r.ok){sync.textContent=\'Actualizado \'+new Date().toLocaleTimeString(\'es-UY\',{timeZone:\'America/Montevideo\'});}}catch(e){sync.textContent=\'Sin conexión\';}},30000);\n{% if current_user.role==\'seller\' %}\nlet watchId=null; const status=document.getElementById(\'trackingStatus\');\nasync function sendPosition(p){let battery=null;try{if(navigator.getBattery){battery=(await navigator.getBattery()).level*100}}catch(e){} const body={latitude:p.coords.latitude,longitude:p.coords.longitude,accuracy:p.coords.accuracy,speed:p.coords.speed,heading:p.coords.heading,battery};const r=await fetch(\'/api/location\',{method:\'POST\',headers:{\'Content-Type\':\'application/json\'},body:JSON.stringify(body)});if(!r.ok)throw new Error(\'No se pudo sincronizar\');status.textContent=\'Ubicación enviada: \'+new Date().toLocaleTimeString(\'es-UY\',{timeZone:\'America/Montevideo\'})+\' · precisión \'+Math.round(p.coords.accuracy)+\' m\';}\nfunction startTracking(){if(!navigator.geolocation){status.textContent=\'Este dispositivo no ofrece geolocalización.\';return}if(watchId!==null)return;localStorage.setItem(\'rimelTracking\',\'1\');status.textContent=\'Solicitando permiso de ubicación…\';watchId=navigator.geolocation.watchPosition(sendPosition,e=>status.textContent=\'GPS: \'+e.message,{enableHighAccuracy:true,maximumAge:15000,timeout:20000});}\nfunction stopTracking(){if(watchId!==null)navigator.geolocation.clearWatch(watchId);watchId=null;localStorage.removeItem(\'rimelTracking\');status.textContent=\'Seguimiento detenido.\';}\ndocument.getElementById(\'startTracking\').onclick=startTracking;document.getElementById(\'stopTracking\').onclick=stopTracking;if(localStorage.getItem(\'rimelTracking\')===\'1\')startTracking();\n{% endif %}\n</script>{% endblock %}\n', 'history.html': '{% extends \'base.html\' %}{% block content %}\n<div class="section-head"><div><h1>Historial de visitas</h1><p class="muted">Fechas y horas mostradas en la zona horaria de Rivera, Uruguay.</p></div><a class="button excel" href="/export/history.xlsx?{{ request.query_string.decode() }}">Exportar a Excel</a></div>\n<form class="filters" method="get">\n{% if current_user.role==\'manager\' %}<select name="seller"><option value="">Todos los vendedores</option>{% for s in sellers %}<option value="{{s}}" {{\'selected\' if s==selected_seller}}>{{s}}</option>{% endfor %}</select>{% endif %}\n<select name="result"><option value="">Todos los resultados</option>{% for r in allowed_results %}<option value="{{r}}" {{\'selected\' if r==selected_result}}>{{r}}</option>{% endfor %}</select>\n<label class="compact">Desde<input type="date" name="date_from" value="{{date_from}}"></label><label class="compact">Hasta<input type="date" name="date_to" value="{{date_to}}"></label><button>Aplicar</button><a class="button secondary" href="/history">Limpiar</a>\n</form>\n<section class="panel"><div class="table-wrap"><table><thead><tr><th>Fecha y hora Rivera</th><th>Vendedor</th><th>Cliente</th><th>Resultado</th><th>Observaciones</th><th>GPS</th></tr></thead><tbody>{% for v in visits %}<tr><td>{{v.created_at|rivera_dt}}</td><td>{{v.seller}}</td><td><b>{{v.client.name}}</b><small>{{v.client.code}}</small></td><td>{{v.result}}</td><td>{{v.notes or \'\'}}</td><td>{{\'Sí\' if v.latitude is not none and v.longitude is not none else \'No\'}}</td></tr>{% else %}<tr><td colspan="6">No hay registros para los filtros seleccionados.</td></tr>{% endfor %}</tbody></table></div></section>\n{% endblock %}\n', 'login.html': '{% extends \'base.html\' %}{% block title %}Acceso RIMEL{% endblock %}{% block content %}<section class="login-card"><h1>Ingresar</h1><form method="post"><label>Usuario<input name="username" required autocomplete="username"></label><label>Contraseña<input type="password" name="password" required autocomplete="current-password"></label><button>Ingresar</button></form><p class="muted">La aplicación puede instalarse desde el menú del navegador: “Agregar a pantalla de inicio”.</p></section>{% endblock %}\n', 'tracking.html': '{% extends \'base.html\' %}{% block content %}\n<div class="topline"><div><h1>Ubicación de vendedores</h1><p class="muted">Última posición enviada durante el seguimiento activo. Horarios de Rivera, Uruguay.</p></div><div class="actions"><button type="button" class="secondary" id="centerMap">Centrar mapa</button><span id="updated" class="pill">Actualizando…</span></div></div>\n<section class="panel map-panel"><div id="map" class="live-map"></div></section>\n<section class="panel"><h2>Estado de dispositivos</h2><div class="table-wrap small-table"><table><thead><tr><th>Vendedor</th><th>Estado</th><th>Última actualización</th><th>Precisión</th><th>Velocidad</th><th>Mapa</th></tr></thead><tbody id="locationRows"></tbody></table></div></section>\n{% endblock %}{% block scripts %}\n<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">\n<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>\n<script>\nlet map, layer, markers={}, accuracyCircles={}, lastBounds=[];\nfunction ago(seconds){if(seconds<60)return \'hace \'+seconds+\' s\';if(seconds<3600)return \'hace \'+Math.floor(seconds/60)+\' min\';return \'hace \'+Math.floor(seconds/3600)+\' h\'}\nfunction initMap(){map=L.map(\'map\',{zoomControl:true,preferCanvas:true}).setView([-30.905,-55.55],13);layer=L.tileLayer(\'https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png\',{subdomains:\'abc\',maxZoom:19,attribution:\'&copy; OpenStreetMap\'}).addTo(map);setTimeout(()=>map.invalidateSize(true),200);window.addEventListener(\'resize\',()=>map.invalidateSize(false));}\nfunction centerAll(){map.invalidateSize(true);if(lastBounds.length===1)map.setView(lastBounds[0],16);else if(lastBounds.length>1)map.fitBounds(lastBounds,{padding:[45,45],maxZoom:16});else map.setView([-30.905,-55.55],13)}\nasync function refresh(){try{const r=await fetch(\'/api/locations\',{cache:\'no-store\'});if(!r.ok)throw new Error(\'Error \'+r.status);const data=await r.json();const rows=document.getElementById(\'locationRows\');rows.innerHTML=\'\';lastBounds=[];const active=new Set();for(const x of data){const tr=document.createElement(\'tr\');if(x.latitude!==undefined){active.add(x.seller);const state=x.online?\'<span class="status online">En línea</span>\':\'<span class="status offline">Sin actualización reciente</span>\';tr.innerHTML=`<td><b>${x.seller}</b></td><td>${state}</td><td>${x.local_time||ago(x.age_seconds)}<small>${ago(x.age_seconds)}</small></td><td>${x.accuracy!=null?Math.round(x.accuracy)+\' m\':\'—\'}</td><td>${x.speed!=null?Math.round(x.speed*3.6)+\' km/h\':\'—\'}</td><td><a class="button secondary" target="_blank" href="https://www.google.com/maps/search/?api=1&query=${x.latitude},${x.longitude}">Abrir</a></td>`;const pos=[x.latitude,x.longitude];if(!markers[x.seller])markers[x.seller]=L.marker(pos).addTo(map);else markers[x.seller].setLatLng(pos);markers[x.seller].bindPopup(`<b>${x.seller}</b><br>${x.local_time||\'\'}<br>Precisión: ${x.accuracy?Math.round(x.accuracy)+\' m\':\'—\'}`);if(x.accuracy){if(!accuracyCircles[x.seller])accuracyCircles[x.seller]=L.circle(pos,{radius:x.accuracy,weight:1,fillOpacity:.08}).addTo(map);else accuracyCircles[x.seller].setLatLng(pos).setRadius(x.accuracy);}lastBounds.push(pos);}else{tr.innerHTML=`<td><b>${x.seller}</b></td><td><span class="status offline">Sin datos</span></td><td>—</td><td>—</td><td>—</td><td>—</td>`;}rows.appendChild(tr);}for(const name of Object.keys(markers)){if(!active.has(name)){map.removeLayer(markers[name]);delete markers[name];if(accuracyCircles[name]){map.removeLayer(accuracyCircles[name]);delete accuracyCircles[name];}}}document.getElementById(\'updated\').textContent=\'Actualizado \'+new Date().toLocaleTimeString(\'es-UY\',{timeZone:\'America/Montevideo\'});map.invalidateSize(false);}catch(e){document.getElementById(\'updated\').textContent=\'Error al actualizar\';console.error(e)}}\ninitMap();document.getElementById(\'centerMap\').onclick=centerAll;refresh().then(centerAll);setInterval(refresh,20000);setTimeout(centerAll,600);\n</script>{% endblock %}\n', 'users.html': '{% extends \'base.html\' %}{% block content %}<h1>Usuarios y accesos</h1><div class="grid2"><section class="panel"><h2>Crear usuario</h2><form method="post"><label>Nombre completo<input name="full_name" required></label><label>Usuario<input name="username" required></label><label>Contraseña inicial<input name="password" type="password" minlength="8" required></label><label>Perfil<select name="role"><option value="seller">Vendedor</option><option value="manager">Gerencial</option></select></label><label>Vendedor asociado<select name="seller_name"><option value="">No corresponde</option>{% for s in sellers %}<option>{{s}}</option>{% endfor %}</select></label><button>Crear usuario</button></form></section><section class="panel"><h2>Usuarios existentes</h2><div class="table-wrap"><table><thead><tr><th>Usuario</th><th>Nombre</th><th>Perfil</th><th>Vendedor</th><th>Restablecer</th></tr></thead><tbody>{% for u in users %}<tr><td>{{u.username}}</td><td>{{u.full_name}}</td><td>{{u.role}}</td><td>{{u.seller_name or \'-\'}}</td><td><form method="post" action="/users/{{u.id}}/reset" class="inline"><input name="password" value="Rimel2026!" minlength="8"><button>Restablecer</button></form></td></tr>{% endfor %}</tbody></table></div></section></div>{% endblock %}\n'}
app.jinja_loader = DictLoader(EMBEDDED_TEMPLATES)

ALLOWED_RESULTS = [
    'PEDIDO CONCRETADO', 'SIN NECESIDAD DE REPOSICION', 'CLIENTE AUSENTE',
    'CLIENTE CERRADO', 'VISITA REPROGRAMADA', 'CLIENTE INACTIVO'
]
DAYS = ['LUNES','MARTES','MIERCOLES','JUEVES','VIERNES','SABADO','DOMINGO']
APP_VERSION = '2026.07.22.13'
RIVERA_TZ = ZoneInfo('America/Montevideo')

def utc_now_naive():
    return datetime.now(timezone.utc).replace(tzinfo=None)

def rivera_now():
    return datetime.now(RIVERA_TZ)

def to_rivera(value):
    if value is None:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(RIVERA_TZ)

def local_day_utc_bounds(day_value):
    local_start = datetime.combine(day_value, datetime.min.time(), tzinfo=RIVERA_TZ)
    local_end = local_start + timedelta(days=1)
    return (local_start.astimezone(timezone.utc).replace(tzinfo=None),
            local_end.astimezone(timezone.utc).replace(tzinfo=None))

def local_range_utc_bounds(date_from, date_to):
    start, _ = local_day_utc_bounds(date_from)
    _, end = local_day_utc_bounds(date_to)
    return start, end

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='seller')
    seller_name = db.Column(db.String(120), nullable=True)
    active = db.Column(db.Boolean, default=True)
    must_change_password = db.Column(db.Boolean, default=True)

class Client(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(30), nullable=False, index=True)
    name = db.Column(db.String(200), nullable=False)
    address = db.Column(db.String(300), nullable=False)
    seller = db.Column(db.String(120), nullable=False, index=True)
    day = db.Column(db.String(20), nullable=False, index=True)
    original_order = db.Column(db.Integer, default=0)
    optimized_order = db.Column(db.Integer, nullable=True)
    latitude = db.Column(db.Float, nullable=True)
    longitude = db.Column(db.Float, nullable=True)
    active = db.Column(db.Boolean, default=True)

class Visit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'), nullable=False)
    seller = db.Column(db.String(120), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    result = db.Column(db.String(80), nullable=False)
    notes = db.Column(db.Text, nullable=True)
    latitude = db.Column(db.Float, nullable=True)
    longitude = db.Column(db.Float, nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now_naive, nullable=False, index=True)
    client = db.relationship('Client')
    user = db.relationship('User')


class LocationPing(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    seller = db.Column(db.String(120), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    latitude = db.Column(db.Float, nullable=False)
    longitude = db.Column(db.Float, nullable=False)
    accuracy = db.Column(db.Float, nullable=True)
    speed = db.Column(db.Float, nullable=True)
    heading = db.Column(db.Float, nullable=True)
    battery = db.Column(db.Float, nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now_naive, nullable=False, index=True)
    user = db.relationship('User')

class Complaint(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    visit_id = db.Column(db.Integer, db.ForeignKey('visit.id'), nullable=False)
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'), nullable=False)
    seller = db.Column(db.String(120), nullable=False, index=True)
    text = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(30), default='PENDIENTE')
    category = db.Column(db.String(60), default='OTRA')
    priority = db.Column(db.String(20), default='MEDIA')
    resolution = db.Column(db.Text, nullable=True)
    resolved_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now_naive, nullable=False)
    visit = db.relationship('Visit')
    client = db.relationship('Client')


def current_user():
    uid = session.get('user_id')
    return db.session.get(User, uid) if uid else None

@app.after_request
def disable_frontend_cache(response):
    if response.mimetype in ('text/html', 'application/json', 'application/javascript'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

@app.context_processor
def inject_globals():
    return {'current_user': current_user(), 'allowed_results': ALLOWED_RESULTS, 'app_version': APP_VERSION}

@app.template_filter('rivera_dt')
def rivera_dt_filter(value, fmt='%d/%m/%Y %H:%M'):
    converted = to_rivera(value)
    return converted.strftime(fmt) if converted else ''

def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user(): return redirect(url_for('login'))
        return fn(*args, **kwargs)
    return wrapper

def manager_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        u=current_user()
        if not u or u.role != 'manager':
            flash('Acceso restringido al perfil gerencial.', 'error')
            return redirect(url_for('dashboard'))
        return fn(*args, **kwargs)
    return wrapper

def normalize_day(v):
    return str(v or '').strip().upper().replace('É','E').replace('Á','A').replace('Í','I').replace('Ó','O').replace('Ú','U')


def ensure_schema_updates():
    """Aplica cambios de esquema simples sin borrar la base existente."""
    inspector = inspect(db.engine)
    if 'client' not in inspector.get_table_names():
        return
    statements = []
    client_columns = {c['name'] for c in inspector.get_columns('client')}
    if 'latitude' not in client_columns:
        statements.append('ALTER TABLE client ADD COLUMN latitude FLOAT')
    if 'longitude' not in client_columns:
        statements.append('ALTER TABLE client ADD COLUMN longitude FLOAT')
    if 'complaint' in inspector.get_table_names():
        complaint_columns = {c['name'] for c in inspector.get_columns('complaint')}
        if 'category' not in complaint_columns:
            statements.append("ALTER TABLE complaint ADD COLUMN category VARCHAR(60) DEFAULT 'OTRA'")
        if 'priority' not in complaint_columns:
            statements.append("ALTER TABLE complaint ADD COLUMN priority VARCHAR(20) DEFAULT 'MEDIA'")
        if 'resolution' not in complaint_columns:
            statements.append('ALTER TABLE complaint ADD COLUMN resolution TEXT')
        if 'resolved_at' not in complaint_columns:
            statements.append('ALTER TABLE complaint ADD COLUMN resolved_at TIMESTAMP')
    for statement in statements:
        db.session.execute(text(statement))
    if statements:
        db.session.commit()

def route_order_expression():
    return case({day: idx for idx, day in enumerate(DAYS, start=1)}, value=Client.day, else_=99)

def next_client_code():
    max_id = db.session.query(db.func.max(Client.id)).scalar() or 0
    return f'CLI-{max_id + 1:06d}'

def next_route_order(seller, day):
    current = db.session.query(db.func.max(db.func.coalesce(Client.optimized_order, Client.original_order))).filter(
        Client.seller == seller, Client.day == day
    ).scalar() or 0
    return int(current) + 1

def apply_embedded_route_orders():
    """Aplica el orden diario precargado sin eliminar clientes, visitas ni historial."""
    try:
        with open(BASE/'routes_embedded.json', encoding='utf-8') as f:
            rows=json.load(f).get('routes', [])
    except Exception:
        return
    by_code={str(r.get('provisional_code','')).strip().upper():r for r in rows}
    by_key={(str(r.get('seller','')).strip().title(),normalize_day(r.get('day')),str(r.get('name','')).strip().upper(),str(r.get('address','')).strip().upper()):r for r in rows}
    changed=0
    for c in Client.query.all():
        r=by_code.get(str(c.code or '').strip().upper()) or by_key.get((str(c.seller or '').strip().title(),normalize_day(c.day),str(c.name or '').strip().upper(),str(c.address or '').strip().upper()))
        if not r: continue
        order=int(r.get('approved_order') or 0)
        if order and c.optimized_order!=order:
            c.optimized_order=order;changed+=1
    if changed: db.session.commit()

def seed_data():
    if Client.query.count() == 0:
        with open(BASE/'routes_embedded.json', encoding='utf-8') as f: data=json.load(f)
        for r in data['routes']:
            db.session.add(Client(code=r['provisional_code'], name=r['name'], address=r['address'], seller=r['seller'].title(), day=normalize_day(r['day']), original_order=r.get('approved_order') or 0, optimized_order=r.get('approved_order') or 0))
    if User.query.count() == 0:
        defaults = [
            ('gerencia','Gerencia RIMEL','manager',None,'Rimel2026!'),
            ('gerson','Gerson','seller','Gerson','Gerson2026!'),
            ('eduardo','Eduardo','seller','Eduardo','Eduardo2026!'),
            ('victoria','Victoria','seller','Victoria','Victoria2026!')]
        for username,full,role,seller,pw in defaults:
            db.session.add(User(username=username,password_hash=generate_password_hash(pw),full_name=full,role=role,seller_name=seller,must_change_password=False))
    db.session.commit()

@app.before_request
def ensure_db():
    if not getattr(app, '_initialized', False):
        db.create_all(); ensure_schema_updates(); seed_data(); apply_embedded_route_orders(); app._initialized=True

@app.get('/health')
def health():
    return {'status':'ok','version':APP_VERSION,'timezone':'America/Montevideo','server_time_rivera':rivera_now().isoformat()}

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method=='POST':
        u=User.query.filter_by(username=request.form.get('username','').strip().lower(), active=True).first()
        if not u or not check_password_hash(u.password_hash, request.form.get('password','')):
            flash('Usuario o contraseña incorrectos.', 'error')
        else:
            session.clear(); session['user_id']=u.id
            return redirect(url_for('change_password') if u.must_change_password else url_for('dashboard'))
    return render_template('login.html')

@app.get('/logout')
def logout(): session.clear(); return redirect(url_for('login'))

@app.route('/change-password', methods=['GET','POST'])
@login_required
def change_password():
    u=current_user()
    if request.method=='POST':
        p=request.form.get('password',''); c=request.form.get('confirm','')
        if len(p)<8: flash('La contraseña debe tener al menos 8 caracteres.','error')
        elif p!=c: flash('Las contraseñas no coinciden.','error')
        else:
            u.password_hash=generate_password_hash(p); u.must_change_password=False; db.session.commit(); flash('Contraseña actualizada.','success'); return redirect(url_for('dashboard'))
    return render_template('change_password.html')

@app.get('/')
@login_required
def dashboard():
    u=current_user()
    seller_filter=request.args.get('seller','').strip()
    day_filter=normalize_day(request.args.get('day',''))
    if u.role=='seller' and not day_filter:
        weekday_idx=rivera_now().weekday()
        if weekday_idx < 6: day_filter=DAYS[weekday_idx]
    date_from_raw=request.args.get('date_from','').strip()
    date_to_raw=request.args.get('date_to','').strip()
    try: date_from=date.fromisoformat(date_from_raw) if date_from_raw else date.today()
    except ValueError: date_from=date.today()
    try: date_to=date.fromisoformat(date_to_raw) if date_to_raw else date.today()
    except ValueError: date_to=date.today()
    if date_to < date_from: date_from, date_to = date_to, date_from
    start_dt, end_dt = local_range_utc_bounds(date_from, date_to)

    cq=Client.query.filter_by(active=True)
    if u.role=='seller':
        seller_filter=u.seller_name
        cq=cq.filter_by(seller=u.seller_name)
    elif seller_filter:
        cq=cq.filter_by(seller=seller_filter)
    if day_filter: cq=cq.filter_by(day=day_filter)
    clients=cq.order_by(route_order_expression(), db.func.coalesce(Client.optimized_order, Client.original_order), Client.name).all()
    client_ids=[c.id for c in clients]

    vq=Visit.query.filter(Visit.created_at>=start_dt, Visit.created_at<end_dt)
    if u.role=='seller': vq=vq.filter_by(seller=u.seller_name)
    elif seller_filter: vq=vq.filter_by(seller=seller_filter)
    if day_filter and client_ids: vq=vq.filter(Visit.client_id.in_(client_ids))
    elif day_filter and not client_ids: vq=vq.filter(db.text('1=0'))
    period_visits=vq.all()

    total=len(clients)
    visits=len(period_visits)
    unique=len({v.client_id for v in period_visits})
    orders=sum(v.result=='PEDIDO CONCRETADO' for v in period_visits)
    no_restock=sum(v.result=='SIN NECESIDAD DE REPOSICION' for v in period_visits)
    absent=sum(v.result=='CLIENTE AUSENTE' for v in period_visits)
    closed=sum(v.result=='CLIENTE CERRADO' for v in period_visits)
    reprogrammed=sum(v.result=='VISITA REPROGRAMADA' for v in period_visits)
    inactive=sum(v.result=='CLIENTE INACTIVO' for v in period_visits)
    effective=orders+no_restock
    contacted=effective+reprogrammed
    gps_count=sum(v.latitude is not None and v.longitude is not None for v in period_visits)
    notes_count=sum(bool(v.notes) for v in period_visits)
    repeat_visits=max(visits-unique,0)
    period_days=max((date_to-date_from).days+1,1)

    complaints_q=Complaint.query.filter(Complaint.created_at>=start_dt, Complaint.created_at<end_dt)
    if u.role=='seller': complaints_q=complaints_q.filter_by(seller=u.seller_name)
    elif seller_filter: complaints_q=complaints_q.filter_by(seller=seller_filter)
    complaints_total=complaints_q.count()
    complaints_pending=complaints_q.filter_by(status='PENDIENTE').count()

    active_sellers_q=db.session.query(Visit.seller).filter(Visit.created_at>=start_dt, Visit.created_at<end_dt).distinct()
    if seller_filter: active_sellers_q=active_sellers_q.filter(Visit.seller==seller_filter)
    active_sellers=active_sellers_q.count()

    metrics={
      'planned':total,'visits':visits,'unique':unique,
      'compliance':round(unique/total*100,1) if total else 0,
      'orders':orders,'conversion':round(orders/effective*100,1) if effective else 0,
      'contact_rate':round(contacted/visits*100,1) if visits else 0,
      'no_restock':no_restock,'absent':absent,'closed':closed,'reprogrammed':reprogrammed,'inactive':inactive,
      'gps_rate':round(gps_count/visits*100,1) if visits else 0,
      'notes_rate':round(notes_count/visits*100,1) if visits else 0,
      'repeat_visits':repeat_visits,'avg_daily':round(visits/period_days,1),
      'complaints':complaints_total,'complaints_pending':complaints_pending,
      'complaint_rate':round(complaints_total/visits*100,1) if visits else 0,
      'active_sellers':active_sellers
    }
    sellers=sorted([x[0] for x in db.session.query(Client.seller).distinct().all()])

    seller_stats=[]
    if u.role=='manager':
        for seller in sellers:
            sq=Visit.query.filter(Visit.seller==seller, Visit.created_at>=start_dt, Visit.created_at<end_dt)
            sv=sq.all(); su=len({v.client_id for v in sv}); so=sum(v.result=='PEDIDO CONCRETADO' for v in sv)
            se=sum(v.result in ('PEDIDO CONCRETADO','SIN NECESIDAD DE REPOSICION') for v in sv)
            sp=Client.query.filter_by(active=True,seller=seller)
            if day_filter: sp=sp.filter_by(day=day_filter)
            planned=sp.count()
            latest=LocationPing.query.filter_by(seller=seller).order_by(LocationPing.created_at.desc()).first()
            seller_stats.append({'seller':seller,'planned':planned,'visits':len(sv),'unique':su,
              'compliance':round(su/planned*100,1) if planned else 0,'orders':so,
              'conversion':round(so/se*100,1) if se else 0,'last_seen':latest.created_at if latest else None})

    route_map_clients=[{'id':c.id,'code':c.code,'name':c.name,'address':c.address,'seller':c.seller,'day':c.day,'latitude':c.latitude,'longitude':c.longitude,'order':c.optimized_order or c.original_order} for c in clients]
    return render_template('dashboard.html', clients=clients, metrics=metrics, sellers=sellers,
      selected_seller=seller_filter if u.role=='manager' else '', selected_day=day_filter, route_map_clients=route_map_clients,
      date_from=date_from.isoformat(),date_to=date_to.isoformat(),seller_stats=seller_stats)

@app.post('/visits')
@login_required
def create_visit():
    u=current_user(); client=db.session.get(Client, int(request.form['client_id']))
    if not client or (u.role=='seller' and client.seller!=u.seller_name): return ('No autorizado',403)
    result=request.form.get('result','').strip().upper()
    if result not in ALLOWED_RESULTS: flash('Resultado inválido.','error'); return redirect(url_for('dashboard'))
    def fnum(v):
        try:return float(v)
        except:return None
    visit=Visit(client_id=client.id,seller=client.seller,user_id=u.id,result=result,notes=request.form.get('notes','').strip() or None,latitude=fnum(request.form.get('latitude')),longitude=fnum(request.form.get('longitude')))
    db.session.add(visit); db.session.flush()
    complaint=request.form.get('complaint','').strip()
    if complaint:
        db.session.add(Complaint(visit_id=visit.id, client_id=client.id, seller=client.seller, text=complaint,
                                 category=request.form.get('complaint_category','OTRA'),
                                 priority=request.form.get('complaint_priority','MEDIA')))
    db.session.commit(); flash('Visita registrada y sincronizada con Gerencia.','success'); return redirect(url_for('dashboard'))

@app.get('/history')
@login_required
def history():
    u = current_user()
    seller = request.args.get('seller','').strip()
    result = request.args.get('result','').strip()
    date_from_raw = request.args.get('date_from','').strip()
    date_to_raw = request.args.get('date_to','').strip()
    q = Visit.query
    if u.role == 'seller':
        seller = u.seller_name
        q = q.filter_by(seller=u.seller_name)
    elif seller:
        q = q.filter_by(seller=seller)
    if result:
        q = q.filter_by(result=result)
    if date_from_raw:
        try:
            start_dt, _ = local_day_utc_bounds(date.fromisoformat(date_from_raw)); q=q.filter(Visit.created_at>=start_dt)
        except ValueError: pass
    if date_to_raw:
        try:
            _, end_dt = local_day_utc_bounds(date.fromisoformat(date_to_raw)); q=q.filter(Visit.created_at<end_dt)
        except ValueError: pass
    visits=q.order_by(Visit.created_at.desc()).limit(2000).all()
    sellers=sorted([x[0] for x in db.session.query(Visit.seller).distinct().all()])
    return render_template('history.html', visits=visits, sellers=sellers, selected_seller=seller,
                           selected_result=result, date_from=date_from_raw, date_to=date_to_raw)

@app.get('/complaints')
@login_required
def complaints():
    u=current_user(); seller=request.args.get('seller','').strip(); status=request.args.get('status','').strip()
    priority=request.args.get('priority','').strip(); q=Complaint.query
    if u.role=='seller': seller=u.seller_name; q=q.filter_by(seller=u.seller_name)
    elif seller: q=q.filter_by(seller=seller)
    if status: q=q.filter_by(status=status)
    if priority: q=q.filter_by(priority=priority)
    rows=q.order_by(Complaint.created_at.desc()).limit(2000).all()
    sellers=sorted([x[0] for x in db.session.query(Complaint.seller).distinct().all()])
    return render_template('complaints.html', complaints=rows, sellers=sellers,
                           selected_seller=seller, selected_status=status, selected_priority=priority)

@app.post('/complaints/<int:cid>/status')
@manager_required
def complaint_status(cid):
    c=db.session.get(Complaint,cid)
    if c:
        c.status=request.form.get('status','PENDIENTE')
        c.category=request.form.get('category',c.category or 'OTRA')
        c.priority=request.form.get('priority',c.priority or 'MEDIA')
        c.resolution=request.form.get('resolution','').strip() or None
        c.resolved_at=utc_now_naive() if c.status=='RESUELTA' else None
        db.session.commit()
    return redirect(url_for('complaints'))


def excel_response(filename, sheet_title, headers, rows):
    wb=Workbook(); ws=wb.active; ws.title=sheet_title[:31]
    ws.append(headers)
    fill=PatternFill('solid', fgColor='0876A8')
    for cell in ws[1]:
        cell.font=Font(color='FFFFFF',bold=True); cell.fill=fill; cell.alignment=Alignment(horizontal='center')
    for row in rows: ws.append(row)
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    for col in range(1, ws.max_column+1):
        max_len=max(len(str(ws.cell(r,col).value or '')) for r in range(1, min(ws.max_row,1000)+1))
        ws.column_dimensions[get_column_letter(col)].width=min(max(max_len+2,10),55)
    output=BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/export/routes.xlsx')
@login_required
def export_routes():
    u=current_user(); seller=request.args.get('seller','').strip(); day_filter=normalize_day(request.args.get('day',''))
    q=Client.query.filter_by(active=True)
    if u.role=='seller': q=q.filter_by(seller=u.seller_name)
    elif seller: q=q.filter_by(seller=seller)
    if day_filter: q=q.filter_by(day=day_filter)
    clients=q.order_by(Client.seller,route_order_expression(),db.func.coalesce(Client.optimized_order,Client.original_order)).all()
    rows=[[c.seller,c.day,c.optimized_order or c.original_order,c.code,c.name,c.address,c.latitude,c.longitude] for c in clients]
    return excel_response('rutas_rimel.xlsx','Rutas',['Vendedor','Día','Orden','Código','Cliente','Dirección','Latitud','Longitud'],rows)

@app.get('/export/history.xlsx')
@login_required
def export_history():
    u=current_user(); seller=request.args.get('seller','').strip(); result=request.args.get('result','').strip()
    q=Visit.query
    if u.role=='seller': q=q.filter_by(seller=u.seller_name)
    elif seller: q=q.filter_by(seller=seller)
    if result: q=q.filter_by(result=result)
    df=request.args.get('date_from','').strip(); dt=request.args.get('date_to','').strip()
    if df:
        try: start,_=local_day_utc_bounds(date.fromisoformat(df)); q=q.filter(Visit.created_at>=start)
        except ValueError: pass
    if dt:
        try: _,end=local_day_utc_bounds(date.fromisoformat(dt)); q=q.filter(Visit.created_at<end)
        except ValueError: pass
    visits=q.order_by(Visit.created_at.desc()).all()
    rows=[]
    for v in visits:
        local=to_rivera(v.created_at)
        rows.append([v.id,local.strftime('%d/%m/%Y') if local else '',local.strftime('%H:%M:%S') if local else '',v.seller,v.client.code,v.client.name,v.client.address,v.result,v.notes or '',v.latitude,v.longitude])
    return excel_response('historial_visitas_rimel.xlsx','Historial',['ID','Fecha Rivera','Hora Rivera','Vendedor','Código','Cliente','Dirección','Resultado','Observaciones','Latitud','Longitud'],rows)

@app.get('/export/complaints.xlsx')
@login_required
def export_complaints():
    u=current_user(); seller=request.args.get('seller','').strip(); status=request.args.get('status','').strip(); priority=request.args.get('priority','').strip()
    q=Complaint.query
    if u.role=='seller': q=q.filter_by(seller=u.seller_name)
    elif seller: q=q.filter_by(seller=seller)
    if status: q=q.filter_by(status=status)
    if priority: q=q.filter_by(priority=priority)
    rows=[]
    for c in q.order_by(Complaint.created_at.desc()).all():
        created=to_rivera(c.created_at); resolved=to_rivera(c.resolved_at)
        rows.append([c.id,created.strftime('%d/%m/%Y %H:%M') if created else '',c.seller,c.client.code,c.client.name,c.category or 'OTRA',c.priority or 'MEDIA',c.text,c.status,c.resolution or '',resolved.strftime('%d/%m/%Y %H:%M') if resolved else ''])
    return excel_response('quejas_rimel.xlsx','Quejas',['ID','Fecha y hora Rivera','Vendedor','Código','Cliente','Categoría','Prioridad','Queja','Estado','Resolución','Cierre'],rows)


@app.get('/clients')
@manager_required
def clients_admin():
    seller = request.args.get('seller', '').strip()
    day = normalize_day(request.args.get('day', ''))
    status = request.args.get('status', 'active').strip()
    search = request.args.get('q', '').strip()
    q = Client.query
    if seller:
        q = q.filter(Client.seller == seller)
    if day:
        q = q.filter(Client.day == day)
    if status == 'active':
        q = q.filter(Client.active.is_(True))
    elif status == 'inactive':
        q = q.filter(Client.active.is_(False))
    if search:
        like = f'%{search}%'
        q = q.filter(or_(Client.name.ilike(like), Client.address.ilike(like), Client.code.ilike(like)))
    rows = q.order_by(route_order_expression(), Client.seller, db.func.coalesce(Client.optimized_order, Client.original_order), Client.name).all()
    sellers = sorted([x[0] for x in db.session.query(Client.seller).distinct().all()])
    return render_template('clients.html', clients=rows, sellers=sellers, selected_seller=seller,
                           selected_day=day, selected_status=status, search=search)

@app.post('/clients/create')
@manager_required
def client_create():
    seller = request.form.get('seller', '').strip().title()
    day = normalize_day(request.form.get('day', ''))
    name = request.form.get('name', '').strip()
    address = request.form.get('address', '').strip()
    code = request.form.get('code', '').strip().upper() or next_client_code()
    if not seller or day not in DAYS or not name or not address:
        flash('Completa vendedor, día, nombre y dirección.', 'error')
        return redirect(url_for('clients_admin'))
    if Client.query.filter_by(code=code).first():
        flash('Ya existe un cliente con ese código.', 'error')
        return redirect(url_for('clients_admin'))
    order = next_route_order(seller, day)
    c = Client(code=code, name=name, address=address, seller=seller, day=day,
               original_order=order, optimized_order=order, active=True)
    db.session.add(c); db.session.commit()
    flash('Cliente agregado. Quedó al final de la ruta; conviene volver a optimizar ese día.', 'success')
    return redirect(url_for('clients_admin', seller=seller, day=day))

@app.route('/clients/<int:client_id>/edit', methods=['GET', 'POST'])
@manager_required
def client_edit(client_id):
    c = db.session.get(Client, client_id)
    if not c:
        return ('Cliente no encontrado', 404)
    if request.method == 'POST':
        seller = request.form.get('seller', '').strip().title()
        day = normalize_day(request.form.get('day', ''))
        name = request.form.get('name', '').strip()
        address = request.form.get('address', '').strip()
        code = request.form.get('code', '').strip().upper()
        if not seller or day not in DAYS or not name or not address or not code:
            flash('Todos los campos son obligatorios.', 'error')
        elif Client.query.filter(Client.code == code, Client.id != c.id).first():
            flash('Ese código ya pertenece a otro cliente.', 'error')
        else:
            route_changed = c.seller != seller or c.day != day or c.address != address
            c.code, c.name, c.address, c.seller, c.day = code, name, address, seller, day
            if route_changed:
                c.optimized_order = next_route_order(seller, day)
                c.latitude = None; c.longitude = None
            db.session.commit()
            flash('Cliente actualizado.' + (' Reoptimiza la ruta modificada.' if route_changed else ''), 'success')
            return redirect(url_for('clients_admin', seller=seller, day=day))
    sellers = sorted([x[0] for x in db.session.query(Client.seller).distinct().all()])
    return render_template('client_edit.html', client=c, sellers=sellers)

@app.post('/clients/<int:client_id>/toggle')
@manager_required
def client_toggle(client_id):
    c = db.session.get(Client, client_id)
    if c:
        c.active = not c.active
        db.session.commit()
        flash('Cliente activado.' if c.active else 'Cliente dado de baja. Su historial se conserva.', 'success')
    return redirect(request.referrer or url_for('clients_admin'))

@app.get('/api/routes/<seller>/<day>')
@manager_required
def route_clients_api(seller, day):
    day = normalize_day(day)
    rows = Client.query.filter_by(seller=seller, day=day, active=True).order_by(
        db.func.coalesce(Client.optimized_order, Client.original_order), Client.name
    ).all()
    return jsonify([{'id': c.id, 'name': c.name, 'address': c.address, 'latitude': c.latitude,
                     'longitude': c.longitude, 'order': c.optimized_order or c.original_order} for c in rows])

@app.post('/api/routes/save-optimized')
@manager_required
def save_optimized_route():
    data = request.get_json(silent=True) or {}
    seller = str(data.get('seller', '')).strip()
    day = normalize_day(data.get('day', ''))
    ordered = data.get('clients') or []
    if not seller or day not in DAYS or not ordered:
        return jsonify({'error': 'Datos incompletos'}), 400
    valid = {c.id: c for c in Client.query.filter_by(seller=seller, day=day, active=True).all()}
    received_ids = []
    for position, item in enumerate(ordered, start=1):
        try:
            cid = int(item.get('id'))
        except (TypeError, ValueError, AttributeError):
            continue
        c = valid.get(cid)
        if not c:
            continue
        c.optimized_order = position
        received_ids.append(cid)
        try:
            c.latitude = float(item['latitude']) if item.get('latitude') is not None else c.latitude
            c.longitude = float(item['longitude']) if item.get('longitude') is not None else c.longitude
        except (TypeError, ValueError):
            pass
    # Los no geocodificados quedan al final, conservando su orden anterior.
    remaining = [c for cid, c in valid.items() if cid not in received_ids]
    remaining.sort(key=lambda c: (c.optimized_order or c.original_order or 999999, c.name))
    pos = len(received_ids)
    for c in remaining:
        pos += 1; c.optimized_order = pos
    db.session.commit()
    return jsonify({'ok': True, 'saved': len(received_ids), 'total': len(valid)})

@app.post('/api/clients/save-coordinate')
@login_required
def save_client_coordinate():
    u=current_user();data=request.get_json(silent=True) or {}
    try: client_id=int(data.get('client_id'));lat=float(data.get('latitude'));lon=float(data.get('longitude'))
    except (TypeError,ValueError): return jsonify({'error':'Datos inválidos'}),400
    if not (-90<=lat<=90 and -180<=lon<=180): return jsonify({'error':'Coordenadas inválidas'}),400
    c=db.session.get(Client,client_id)
    if not c or (u.role=='seller' and c.seller!=u.seller_name): return jsonify({'error':'No autorizado'}),403
    c.latitude=lat;c.longitude=lon;db.session.commit();return jsonify({'ok':True})

@app.route('/users', methods=['GET','POST'])
@manager_required
def users():
    if request.method=='POST':
        username=request.form['username'].strip().lower()
        if User.query.filter_by(username=username).first(): flash('El usuario ya existe.','error')
        else:
            role=request.form['role']; seller=request.form.get('seller_name','').strip() or None
            db.session.add(User(username=username,full_name=request.form['full_name'].strip(),role=role,seller_name=seller,password_hash=generate_password_hash(request.form['password']),must_change_password=False)); db.session.commit(); flash('Usuario creado.','success')
    sellers=sorted([x[0] for x in db.session.query(Client.seller).distinct().all()])
    return render_template('users.html', users=User.query.order_by(User.role,User.full_name).all(), sellers=sellers)

@app.post('/users/<int:uid>/reset')
@manager_required
def reset_user(uid):
    u=db.session.get(User,uid); pw=request.form.get('password','Rimel2026!')
    if u: u.password_hash=generate_password_hash(pw); u.must_change_password=True; db.session.commit(); flash('Contraseña restablecida.','success')
    return redirect(url_for('users'))

@app.get('/api/summary')
@login_required
def api_summary():
    u=current_user(); since,_=local_day_utc_bounds(rivera_now().date()); q=Visit.query.filter(Visit.created_at>=since)
    if u.role=='seller': q=q.filter_by(seller=u.seller_name)
    rows=q.all(); return jsonify({'visits':len(rows),'orders':sum(v.result=='PEDIDO CONCRETADO' for v in rows),'last_update':rivera_now().isoformat()})


@app.post('/api/location')
@login_required
def save_location():
    u=current_user()
    if u.role!='seller' or not u.seller_name: return jsonify({'error':'Solo vendedores'}),403
    data=request.get_json(silent=True) or {}
    try:
        lat=float(data['latitude']); lon=float(data['longitude'])
    except (KeyError,TypeError,ValueError): return jsonify({'error':'Coordenadas inválidas'}),400
    def val(name):
        try: return float(data[name]) if data.get(name) is not None else None
        except (TypeError,ValueError): return None
    ping=LocationPing(seller=u.seller_name,user_id=u.id,latitude=lat,longitude=lon,
        accuracy=val('accuracy'),speed=val('speed'),heading=val('heading'),battery=val('battery'))
    db.session.add(ping); db.session.commit()
    return jsonify({'ok':True,'saved_at':ping.created_at.isoformat()})

@app.get('/tracking')
@manager_required
def tracking():
    sellers=sorted([x[0] for x in db.session.query(Client.seller).distinct().all()])
    return render_template('tracking.html', sellers=sellers)

@app.get('/api/locations')
@manager_required
def api_locations():
    sellers=[x[0] for x in db.session.query(Client.seller).distinct().all()]
    result=[]
    now=utc_now_naive()
    for seller in sellers:
        p=LocationPing.query.filter_by(seller=seller).order_by(LocationPing.created_at.desc()).first()
        if p:
            age=(now-p.created_at).total_seconds()
            result.append({'seller':seller,'latitude':p.latitude,'longitude':p.longitude,
              'accuracy':p.accuracy,'speed':p.speed,'heading':p.heading,
              'created_at':p.created_at.isoformat()+'Z',
              'local_time':to_rivera(p.created_at).strftime('%d/%m/%Y %H:%M:%S'),
              'age_seconds':round(age),'online':age<=180})
        else:
            result.append({'seller':seller,'online':False})
    return jsonify(result)

@app.get('/client/<int:client_id>/map')
@login_required
def client_map(client_id):
    u=current_user(); c=db.session.get(Client,client_id)
    if not c or (u.role=='seller' and c.seller!=u.seller_name): return ('No autorizado',403)
    return redirect('https://www.google.com/maps/search/?api=1&query='+__import__('urllib.parse').parse.quote_plus(f"{c.address}, Rivera, Uruguay"))

@app.get('/service-worker.js')
def service_worker():
    js = "self.addEventListener('install',e=>self.skipWaiting());self.addEventListener('activate',e=>e.waitUntil(caches.keys().then(k=>Promise.all(k.map(x=>caches.delete(x)))).then(()=>self.registration.unregister()).then(()=>self.clients.claim())));"
    return Response(js, mimetype='application/javascript', headers={'Cache-Control':'no-store, no-cache, must-revalidate, max-age=0'})

if __name__=='__main__':
    with app.app_context(): db.create_all(); ensure_schema_updates(); seed_data()
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT',8000)), debug=False, use_reloader=False)
